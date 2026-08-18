"""Builds an openpyxl Workbook from a reportgen.ReportData. Runs server-side (see views.py) —
never in the browser — so this stays cheap regardless of how many devices/cables a scope
contains; the client only ever downloads the finished .xlsx bytes."""
from __future__ import annotations

from io import BytesIO

from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .reportgen import ReportData

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, 10), 45)


def _border_all(ws, min_row=2):
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            cell.border = BORDER


def _write_cover(ws, data: ReportData):
    ws.title = _("Cover")
    title = data.company_name or _("Cross Journal")
    ws["B2"] = f"{title} — {data.scope_label}"
    ws["B2"].font = Font(size=18, bold=True)
    row = 4
    if data.site_name:
        ws[f"B{row}"] = _("Site:")
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"C{row}"] = data.site_name
        row += 1
    if data.location_name:
        ws[f"B{row}"] = _("Location:")
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"C{row}"] = data.location_name
        row += 1
    ws[f"B{row}"] = {"rack": _("Rack:"), "location": _("Location:"), "site": _("Site:")}[data.scope_kind]
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"C{row}"] = data.scope_label
    row += 2
    ws[f"B{row}"] = _("Generated from NetBox by the Cross Journal plugin.")
    ws[f"B{row}"].font = Font(italic=True, color="888888")
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40


def _write_devices(ws, data: ReportData):
    ws.title = _("Devices")
    headers = [_("Name"), _("Device type"), _("Role"), _("Serial number"), _("Status"),
               _("Location"), _("IP"), _("Tags"), _("Comments")]
    ws.append(headers)
    _style_header(ws)
    for d in data.devices:
        ws.append([d.name, d.device_type, d.role, d.serial, d.status, d.location,
                   d.primary_ip, ", ".join(d.tags), d.comments[:1000]])
    _autosize(ws)
    _border_all(ws)
    ws.freeze_panes = "A2"


def _write_data_cables(ws, data: ReportData):
    ws.title = _("Data")
    headers = [_("Cable"), _("Cable type"), _("Device A"), _("Port A"),
               _("Device B"), _("Port B"), _("Comment"), _("Port comment")]
    ws.append(headers)
    _style_header(ws)
    for c in data.data_cables:
        ws.append([c.label, c.cable_type, c.a_device, c.a_port, c.b_device, c.b_port,
                   c.comments[:1000], c.port_comment[:1000]])
    _autosize(ws)
    _border_all(ws)
    ws.freeze_panes = "A2"


def _write_power_cables(ws, data: ReportData):
    ws.title = _("Power")
    headers = [_("Device"), _("Power port"), _("PDU"), _("Outlet")]
    ws.append(headers)
    _style_header(ws)
    for p in data.power_cables:
        ws.append([p.device, p.port, p.pdu, p.outlet])
    _autosize(ws)
    _border_all(ws)
    ws.freeze_panes = "A2"


def _write_single_sheet(ws, data: ReportData):
    ws.title = _("Cross Journal")
    ws.append([_("DEVICES")])
    ws["A1"].font = Font(bold=True, size=13)
    headers = [_("Name"), _("Device type"), _("Role"), _("Serial number"), _("Status"),
               _("Location"), _("IP"), _("Tags"), _("Comments")]
    ws.append(headers)
    _style_header(ws, row=2)
    for d in data.devices:
        ws.append([d.name, d.device_type, d.role, d.serial, d.status, d.location,
                   d.primary_ip, ", ".join(d.tags), d.comments[:1000]])

    if data.data_cables:
        ws.append([])
        r = ws.max_row + 1
        ws.append([_("DATA")])
        ws.cell(row=r, column=1).font = Font(bold=True, size=13)
        ws.append([_("Cable"), _("Cable type"), _("Device A"), _("Port A"),
                    _("Device B"), _("Port B"), _("Comment"), _("Port comment")])
        _style_header(ws, row=ws.max_row)
        for c in data.data_cables:
            ws.append([c.label, c.cable_type, c.a_device, c.a_port, c.b_device, c.b_port,
                       c.comments[:1000], c.port_comment[:1000]])

    if data.power_cables:
        ws.append([])
        r = ws.max_row + 1
        ws.append([_("POWER")])
        ws.cell(row=r, column=1).font = Font(bold=True, size=13)
        ws.append([_("Device"), _("Power port"), _("PDU"), _("Outlet")])
        _style_header(ws, row=ws.max_row)
        for p in data.power_cables:
            ws.append([p.device, p.port, p.pdu, p.outlet])

    _autosize(ws)


def build_workbook(data: ReportData, layout: str = "split") -> BytesIO:
    wb = Workbook()
    if layout == "single":
        wb.remove(wb.active)
        ws = wb.create_sheet()
        _write_single_sheet(ws, data)
    else:
        _write_cover(wb.active, data)
        _write_devices(wb.create_sheet(), data)
        if data.data_cables:
            _write_data_cables(wb.create_sheet(), data)
        if data.power_cables:
            _write_power_cables(wb.create_sheet(), data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
